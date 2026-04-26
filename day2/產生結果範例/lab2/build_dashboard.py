"""Generate a self-contained HTML dashboard from Financial_Model.xlsx."""
import json
import openpyxl
from pathlib import Path

wb = openpyxl.load_workbook("Financial_Model.xlsx", data_only=True)

def row_by_label(ws, label, col_start=3):
    for row in ws.iter_rows(values_only=True):
        if row[1] == label:
            return list(row)
    return None

# ---- Time axis: years 1..40 ----
years = list(range(1, 41))

# ---- Input Assumptions (base case) ----
ia = wb["Input Assumptions"]
assumptions = {}
for row in ia.iter_rows(values_only=True):
    if row[1] and row[3] is not None and row[2] is not None:
        assumptions[row[1]] = {"unit": row[2], "value": row[3]}

# ---- Operation sheet ----
op = wb["Operation"]
def op_row(label):
    for row in op.iter_rows(values_only=True):
        if row[1] == label:
            # year columns are index 4..43
            return [v if isinstance(v, (int, float)) else 0 for v in row[4:44]]
    return None

traffic_pc = op_row("TRAFFIC - Passenger Car (PC)")
traffic_hv = op_row("TRAFFIC - Heavy Vehicle (HV)")
rev_pc = op_row("REVENUE - Passenger Car (PC)")
rev_hv = op_row("REVENUE - Heavy Vehicle (HV)")
total_rev_nominal = op_row("Total Revenue (nominal)")
maintenance = op_row("Maintenance & SPV costs")

# ---- P&L ----
pl = wb["P&L"]
def pl_row(label):
    for row in pl.iter_rows(values_only=True):
        if row[1] == label:
            return [v if isinstance(v, (int, float)) else 0 for v in row[5:45]]
    return None

ebitda = pl_row("EBITDA")
ebit = pl_row("EBIT")
net_profit = pl_row("Net Profit")
interest_expense = pl_row("Interest")

# ---- Debt ----
dt = wb["Debt"]
def dt_row(label):
    for row in dt.iter_rows(values_only=True):
        if row[1] == label:
            return [v if isinstance(v, (int, float)) else 0 for v in row[5:45]]
    return None

debt_opening = dt_row("Opening Balance")
debt_closing = dt_row("Closing Balance")
principal_rep = dt_row("Principal Repayment")
interest_paid = dt_row("Interests")
debt_service = dt_row("Debt service")
drawdowns = dt_row("Drawdowns")

# ---- CFS ----
cfs = wb["CFS"]
def cfs_row(label):
    for row in cfs.iter_rows(values_only=True):
        if row[1] == label:
            return [v if isinstance(v, (int, float)) else 0 for v in row[5:45]]
    return None

cfads = cfs_row("Cash Flow Available for Debt Service (CFADS)")
dividends = cfs_row("Dividends")
cash_in_hand = cfs_row("Cash in hands")

# ---- Balance Sheet ----
bs = wb["Balance Sheet"]
def bs_row(label):
    for row in bs.iter_rows(values_only=True):
        if row[1] == label:
            return [v if isinstance(v, (int, float)) else 0 for v in row[5:45]]
    return None

asset_bs = bs_row("Asset")
cash_bs = bs_row("Cash in hand")
debt_bs = bs_row("Debt")
equity_bs = bs_row("Equity")

# ---- Ratios ----
rt = wb["Ratios"]
project_irr = None
equity_irr = None
for row in rt.iter_rows(values_only=True):
    if row[1] in ("Project IRR", "Equity IRR"):
        for v in row[3:]:
            if isinstance(v, (int, float)) and 0 < v < 1:
                if row[1] == "Project IRR":
                    project_irr = v
                else:
                    equity_irr = v
                break

# Scalar headline figures
capex = 300000  # k£
total_debt = 214538.18
total_equity = 115520.56
concession_duration = 40
construction_duration = 4
operations_duration = 36

# Sources & uses
arrangement_fee = 3218.07
engagement_fee = 13991.17
capitalized_interest = 12849.50
total_uses = 330058.74

# Peak metrics
peak_revenue = max(total_rev_nominal)
peak_ebitda = max(ebitda)
peak_net_profit = max(net_profit)
total_net_profit = sum(net_profit)
total_dividends = -sum(dividends)
total_cfads = sum(cfads)

data = {
    "years": years,
    "assumptions": assumptions,
    "traffic_pc": traffic_pc,
    "traffic_hv": traffic_hv,
    "rev_pc": rev_pc,
    "rev_hv": rev_hv,
    "total_rev_nominal": total_rev_nominal,
    "maintenance": maintenance,
    "ebitda": ebitda,
    "ebit": ebit,
    "net_profit": net_profit,
    "interest_expense": interest_expense,
    "debt_opening": debt_opening,
    "debt_closing": debt_closing,
    "principal_rep": principal_rep,
    "interest_paid": interest_paid,
    "debt_service": debt_service,
    "drawdowns": drawdowns,
    "cfads": cfads,
    "dividends": dividends,
    "cash_in_hand": cash_in_hand,
    "asset_bs": asset_bs,
    "cash_bs": cash_bs,
    "debt_bs": debt_bs,
    "equity_bs": equity_bs,
    "kpis": {
        "project_irr": project_irr,
        "equity_irr": equity_irr,
        "capex": capex,
        "total_debt": total_debt,
        "total_equity": total_equity,
        "concession_duration": concession_duration,
        "construction_duration": construction_duration,
        "operations_duration": operations_duration,
        "arrangement_fee": arrangement_fee,
        "engagement_fee": engagement_fee,
        "capitalized_interest": capitalized_interest,
        "total_uses": total_uses,
        "peak_revenue": peak_revenue,
        "peak_ebitda": peak_ebitda,
        "peak_net_profit": peak_net_profit,
        "total_net_profit": total_net_profit,
        "total_dividends": total_dividends,
        "total_cfads": total_cfads,
    },
}

Path("data.json").write_text(json.dumps(data, indent=2, default=float))
print("data.json written.")
print(f"Project IRR: {project_irr}")
print(f"Equity IRR: {equity_irr}")
